import os
import numpy as np
import pandas as pd
import openpyxl
import zipfile
import tempfile
import shutil

DATA_PATH = "../data/calce"

# Column indices in the Channel_* sheet (1-indexed from openpyxl)
# Row 1 is the header; data starts at row 2
COL_CYCLE = 6    # Cycle_Index
COL_CURRENT = 7  # Current(A)
COL_VOLTAGE = 8  # Voltage(V)
COL_CHARGE_AH = 9
COL_DISCHARGE_AH = 10


def extract_cell_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = None
    for s in wb.sheetnames:
        if not s.startswith("Channel"):
            continue
        obj = wb[s]
        if hasattr(obj, 'iter_rows'):
            ws = obj
            break
    if ws is None:
        return pd.DataFrame()

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 15:
            continue
        cycle_idx = row[COL_CYCLE - 1]
        current = row[COL_CURRENT - 1]
        voltage = row[COL_VOLTAGE - 1]
        charge_ah = row[COL_CHARGE_AH - 1]
        discharge_ah = row[COL_DISCHARGE_AH - 1]

        if cycle_idx is None:
            continue

        rows.append({
            "cycle": int(cycle_idx),
            "current": float(current) if current is not None else np.nan,
            "voltage": float(voltage) if voltage is not None else np.nan,
            "charge_ah": float(charge_ah) if charge_ah is not None else np.nan,
            "discharge_ah": float(discharge_ah) if discharge_ah is not None else np.nan,
        })

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows)


def extract_cell_cycles(zip_path):
    cell_name = os.path.splitext(os.path.basename(zip_path))[0]
    folder_in_zip = cell_name

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_path, 'r') as z:
            z.extractall(tmpdir)

        cell_dir = os.path.join(tmpdir, folder_in_zip)
        if not os.path.isdir(cell_dir):
            cell_dir = tmpdir

        all_records = []
        cycle_offset = 0
        xlsx_files = sorted([f for f in os.listdir(cell_dir) if f.endswith('.xlsx')])

        for xf in xlsx_files:
            xf_path = os.path.join(cell_dir, xf)
            df = extract_cell_xlsx(xf_path)
            if df.empty:
                continue

            # Per-cycle features within this file using cumulative discharge_ah deltas
            prev_max_dc = 0.0
            file_records = []
            for cyc, g in df.groupby("cycle", sort=True):
                discharge = g[g["current"] < -0.01].copy()
                if discharge.empty:
                    continue

                cur_max_dc = float(discharge["discharge_ah"].max())
                cap_val = cur_max_dc - prev_max_dc
                prev_max_dc = cur_max_dc

                if cap_val <= 0 or not np.isfinite(cap_val):
                    continue

                v_vals = discharge["voltage"].dropna().values
                i_vals = discharge["current"].dropna().values

                if len(v_vals) < 2:
                    continue

                file_records.append({
                    "capacity": cap_val,
                    "avg_voltage": float(np.nanmean(v_vals)),
                    "min_voltage": float(np.nanmin(v_vals)),
                    "avg_current": float(np.nanmean(i_vals)),
                })

            if not file_records:
                continue

            file_df = pd.DataFrame(file_records)
            file_df["cycle"] = range(cycle_offset + 1, cycle_offset + len(file_df) + 1)
            cycle_offset += len(file_df)
            all_records.append(file_df)

    if not all_records:
        return pd.DataFrame()

    cell_df = pd.concat(all_records, ignore_index=True)
    cell_df["avg_temp"] = np.nan
    cell_df["duration"] = np.nan

    initial_cap = cell_df["capacity"].iloc[0]
    if not np.isfinite(initial_cap) or initial_cap <= 0:
        return pd.DataFrame()

    cell_df["SOH"] = cell_df["capacity"] / initial_cap

    eol_idx = cell_df.index[cell_df["SOH"] <= 0.8]
    if len(eol_idx) > 0:
        eol_cycle = int(cell_df.loc[eol_idx[0], "cycle"])
        cell_df["RUL"] = (eol_cycle - cell_df["cycle"]).clip(lower=0)
    else:
        cell_df["RUL"] = cell_df["cycle"].max() - cell_df["cycle"]

    cell_df["cell"] = f"calce_{cell_name}"
    return cell_df


def load_all_calce():
    all_cells = []

    for f in sorted(os.listdir(DATA_PATH)):
        if not f.endswith(".zip"):
            continue
        path = os.path.join(DATA_PATH, f)
        print(f"Loading: {f}")
        try:
            df = extract_cell_cycles(path)
        except Exception as e:
            print(f"  Error: {e}")
            continue
        if df.empty or len(df) < 5:
            print(f"  Skipped: insufficient data")
            continue
        all_cells.append(df)
        print(f"  {df['cell'].iloc[0]}: {len(df)} cycles, "
              f"capacity={df['capacity'].iloc[0]:.3f}Ah, "
              f"SOH range={df['SOH'].min():.3f}-{df['SOH'].max():.3f}")

    if not all_cells:
        raise RuntimeError("No CALCE cells loaded.")

    result = pd.concat(all_cells, ignore_index=True)
    out_csv = "../data/calce_clean.csv"
    result.to_csv(out_csv, index=False)
    print(f"\nSaved: {out_csv} with {len(result)} rows, {result['cell'].nunique()} cells")
    return result


if __name__ == "__main__":
    load_all_calce()
